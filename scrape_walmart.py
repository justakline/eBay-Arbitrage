import requests #Gets the information
import os
import smtplib
from bs4 import BeautifulSoup #Parses through information
from selenium import webdriver 
import time
from WalmartPageInfo import WalmartPageInfo
from selenium.common import exceptions  
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from openpyxl import load_workbook
import traceback
import urllib
import re
from pathlib import Path
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import shutil
from removeEbay import *
from reviseEbay import *

def definePhysicalRows(sheet):
    sheet['A1'] = 'Requirements'
    sheet['A2'] = 'URL'
    sheet['A3'] = 'Title'
    sheet['A4'] = 'Condition'#Always is new
    sheet['A5'] = "Stock"
    sheet['A6'] = 'Brand'
    sheet['A7'] = 'Model'
    sheet['A8'] = 'Maximum Resolution'
    sheet['A9'] = 'Screen Size'
    sheet['A10'] = "Display Technology"
    sheet['A11'] = 'Refresh Rate'
    sheet['A14'] = 'Body'
    sheet['A15'] = 'Price'
    sheet['A16'] = 'Package weight'
    sheet['A17'] = 'Depth'
    sheet['A18'] = 'Width'
    sheet['A19'] = 'Height'
    sheet['A20'] = 'Sell Price'
    sheet['A21'] = 'Total Costs'
    sheet['A22'] = 'Total Profits'
    sheet['A23'] = 'ItemID'
    sheet['A24'] = 'ValueChanged'
    sheet['B24'] = 'False'
    sheet['C1'] = 'Photos'
    sheet['E1'] = 'Features'
    sheet['G1'] = 'Smart TV Features'
    sheet['I1'] = 'Audio/Video Inputs'

#sends an email to me notifying if a price has changed or if we are out of stock
def sendEmail(subject, body):
    server = smtplib.SMTP('64.233.184.108', 587)#connect to google
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.login("justakline@gmail.com", "s5hduV9W")

    message = f"Subject: {subject}\n\n{body}"

    server.sendmail(
        "justakline@gmail.com",
        "justakline@gmail.com",
        message.encode("ascii" , "replace")
    )

    print("mailSent")

#opens a new page and does it in a timely manner
def openNewPage(driver, link):
    ActionChains(driver).context_click(link).key_down(Keys.CONTROL).click(link).perform() #clicks link and opens new tab
    tabs = driver.window_handles #tells tabs to hold onto all of the open tabs in driver
    driver.switch_to_window(tabs[1]) #switches to new tab
    elementToWaitFor = driver.find_element_by_tag_name('h1')
    WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.TAG_NAME, 'h1')))
    page = WalmartPageInfo(driver.current_url, driver)
    return [page,tabs]

#used to look through and find all the prices that need to be changed
#and sends an email
def valueChanged(dirs, every, parentDir):
    for d in dirs:# if this is already in a spreadsheet, then check if the value has changed
        os.chdir(d)
   
        workbook = load_workbook("item.xlsx")
  
        sheetNames = workbook.sheetnames
        sheet = workbook[sheetNames[len(sheetNames)-1]]
        if(every[5] == sheet['B7'].value or every[5] == []):#ie do the model numbers match or does it have one
            if(every[3] == "Out of Stock"): #if we are out of stock
                subject = "URGENT!  " +sheet['B3'].value + " is " + "OUT OF STOCK!!!"
                body = "Change Now!"
                sheet['B5'] = "Out of Stock"
            elif(not sheet['B15'].value == every[13]):#Are the prices equal?
                subject = "Price changed for " + sheet['B3'].value
                body = sheet['B3'].value + " changed  price from " + sheet['B15'].value + " to " + str(every[13]) + "\n change listing price to " #+ every[]
                sheet['B15'] = every[13] #change the price to the current one
                sheet['B24'] = "True"
                # sheet['B20'] = every[]
            print("new = false")
            return True
        workbook.save("item.xlsx")
        workbook.save("eBay.xlsx")
        os.chdir(parentDir)
    return False

def getRidOfAllInvalidDirs(parentDir):
    removeDirs = []
    os.chdir(parentDir)
    listDir = os.listdir()
    print(listDir)
    for dir in listDir:
        os.chdir(dir)
        foundItem = False
        subDir =os.listdir() 
        
        for subFile in subDir:
            if(subFile == "item.xlsx"): #does it have an items section
                foundItem = True
                workbook = load_workbook("item.xlsx")
                sheetNames = workbook.sheetnames
                sheet = workbook[sheetNames[len(sheetNames)-1]]
                if(sheet['B6'].value is None):#does it not have specs
                    foundItem = False
                    print("no specs")
                workbook.close()
            elif(subFile == "images"): #does it have images
                before = os.getcwd()
                os.chdir(subFile)
                foundPictures = (len(os.listdir()) > 0)
                os.chdir(before)
            # os.chdir(subDir)
        os.chdir(parentDir)
        if(not foundItem or not foundPictures):
            removeDirs.append(dir)
            foundItem = False
    os.chdir(parentDir)
    for dir in removeDirs:
        shutil.rmtree(dir, ignore_errors=True)

def addItemDir(row, col, every, images):
    name = re.sub('[\/:*?"<>|]', "", every[1])#directories dont like these characters in its name
    try:
        os.mkdir(name)
    except FileExistsError:
        pass
        
    workbook = Workbook()
    sheetNames = workbook.sheetnames
    rowIndex = 1
    colIndex = 1
    rowcol = str(col[colIndex]) + str(row[rowIndex]) 
    sheet = workbook[sheetNames[len(sheetNames)-1]]
    sheet.title = name #at the title
    numberOfLists = 0
    definePhysicalRows(sheet)
        #walmart
    for thing in every:
        if(isinstance(thing, list)):
            tempRow = 2+ 2*numberOfLists
            tempCol = 1 
            if(tempRow == 4):#This is so that it does not fill in the features section, it does not have that
                tempRow = 6
                numberOfLists += 1
            for i in range(len(thing)):#Numbering each row in the list
                rowcol = str(col[tempRow]) + str(row[tempCol])
                sheet[rowcol] = i
                tempCol += 1
            tempRow = 3 + 2*numberOfLists
            tempCol = 1
            for subThing in thing:
                rowcol = str(col[tempRow]) + str(row[tempCol])
                if(subThing[0:5] == "https"): #if these are images
                    images.append(requests.get(subThing))
                if("," in subThing):
                    features = subThing.split(",")
                    for f in features:
                        sheet[rowcol] = f
                        tempCol+=1
                        rowcol = str(col[tempRow]) + str(row[tempCol])
                else:
                    sheet[rowcol] = subThing
                    tempCol +=1
            rowcol = str(col[colIndex]) + str(row[rowIndex]) 
            numberOfLists += 1
        else:
            sheet[rowcol] = thing
            rowIndex += 1
            rowcol = str(col[colIndex]) + str(row[rowIndex])
    sheet['B23'] = sheet['B7'].value
    os.chdir(name)
    workbook.save(filename = "item.xlsx")
    workbook.save(filename = "eBay.xlsx")
    getImages = True

def addImagesDir(images, parentDir):
    os.mkdir("images")
    os.chdir("images")
    i = 0
    for image in images:#download images
        imageName = "image " + str(i) + ".png"
        file = open(imageName, "wb")
        try:
            file.write(image.content)
        except:
            exceptions.NoSuchAttributeException
        file.close()
        i+=1
    os.chdir(parentDir)

def getAllListedItemLinks(parentDir):
    parentDir = os.getcwd()
    listDir = os.listdir()
    random.shuffle(listDir)
    listed = []
    for dir in listDir:
        os.chdir(dir)
        workbook = load_workbook("item.xlsx")
        sheet = workbook[workbook.sheetnames[0]]
        listed.append(sheet["B3"].value)
        os.chdir(parentDir)
    # os.chdir(parentDir)
    return listed

def scrape_walmart(driver, URL):
    allPages = []


    # options = webdriver.ChromeOptions()
    # options.add_argument(r"user-data-dir= C\Users\justa\AppData\Local\Google\Chrome\User Data\Default")
    # # options.add_argument('headless')
    # options.add_argument('window-size=1200x600')
    # driver = webdriver.Chrome(chrome_options = options)
    try:
        driver.get(URL)
    except:
        exceptions.NoSuchWindowException
        driver.switch_to_window(driver.window_handles[0])
    i = 0
    while(driver.title == "Verify your identity"):
        if (i %1000 == 0):
            print("re")

    the = 40
    row = []
    for i in range(1, 100):
        row.append(i)
    col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 
    'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S' , 'T', 'U', 
    'V', 'W', 'X', 'Y', 'Z', 'AA' , 'AB', 'AD', 'AE' , 'AF', 'AG', 
    'AH' , 'AI', 'AJ', 'AK' , 'AL', 'AM', 'AN' , 'AO', 'AP',
    'AQ' , 'AR', 'AS', 'AT' , 'AU', 'AV', 'AW' , 'AX', 'AY', 'AZ']


    parentDir = os.getcwd()
    print(parentDir)
    if(not ( "allItems" in parentDir)):
        os.chdir("allItems")
        parentDir = os.getcwd()
    
    listed = getAllListedItemLinks(parentDir)

    links = driver.find_elements_by_class_name('product-title-link') #finds all links from product summaries on main page
    random.shuffle(links)
            
   
    for link in links:
        for item in listed: 
            if (link.text.replace("+", "") in item):
                print ("removed " + link.text.replace("+", ""))
                links.remove(link)
                listed.remove(item)
                print()
                break
    print (links)
    linkIndex = 0

    i = 0
    while(linkIndex < len(links)):
        try:
            titlePrice = {}
            missed = 0
            total = 0
            openedNewPage = True
            if(not (os.getcwd() == parentDir)):
                os.chdir(parentDir)
            for link in links:
                openedNewPage = True #We assume that we can open another page
                try:
                    p = openNewPage(driver, link)
                    page = p[0]
                    tabs = p[1]
                except (exceptions.StaleElementReferenceException, exceptions.TimeoutException):
                    print("no link")
                    print(traceback.format_exc())
                    openedNewPage = False
                    if(len(driver.window_handles) >= 2):
                        driver.close()
                        driver.switch_to_window(tabs[0]) #switches back to old tab
                if(openedNewPage):#if have opened a new page
                    every = page.ebayRepresentation()
                    dirs = os.listdir()
                    try:
                        newItem = (not valueChanged(dirs,every, parentDir)) and page.outOfStock == "In Stock"# and page.goodShipping == True
                    except:
                        exceptions.NoSuchElementException
                        newItem = False
                    print("newItem = " + str(newItem))
                    os.chdir(parentDir)
                    if(newItem):#If we have found a new item
                        images = []
                        addItemDir(row, col,every, images)
                        # addImagesDir(images, parentDir)
                    total+=1  
                    driver.close() #closes the new tab
                    driver.switch_to_window(tabs[0]) #switches back to old tab
                    driver.implicitly_wait(5) #Used so that everything has time to load
                os.chdir(parentDir)
                links.remove(link)
        except:
            print(traceback.format_exc())
            print("Something Went Wrong")
            try:
                if(driver.title == "Verify your identity"):
                    sendEmail("ReCaptcha","Wrong")
                    recaptcha = True
                some = 0
                while(recaptcha):
                    if(some % 1000 == 0):
                        print("in Recaptcha")
                    if(not (driver.title == "Verify your identity")):
                        recaptcha = False
                    some+=1
            except: 
                exceptions.UnexpectedAlertPresentException
        linkIndex +=1
    if(len(driver.window_handles) >=2):
        driver.close() #closes the new tab
        driver.switch_to_window(tabs[0]) #switches back to old tab
    getRidOfAllInvalidDirs(parentDir)
    os.chdir(parentDir)
    sendEmail("Finished Scraping", "Good Job!")
    # driver.quit()
        # minute = 60
        # mins = minute * 60
        # time.sleep(mins)

