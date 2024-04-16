import os
import smtplib

from selenium import webdriver 
from openpyxl import Workbook
from openpyxl import load_workbook
from scrape_walmart import *
from ebaysdk.trading import Connection
import shutil
import ebaysdk.trading
from addEbay import *
from removeEbay import *
from reviseEbay import *


# scrape_walmart()
# options  = webdriver.ChromeOptions()
# options.add_argument('window-size=1200x600')
# options.add_argument(r"user-data-dir= C\Users\justa\AppData\Local\Google\Chrome\User Data\Default")
# driver = webdriver.Chrome(chrome_options = options)
os.chdir("allItems")
parentDir = os.getcwd()
listDir = os.listdir()
print(listDir)
for dir in listDir:
    os.chdir(dir)
#     print(dir)
    workbook = load_workbook("item.xlsx")
#     sheet = workbook[workbook.sheetnames[0]]
#     sheet["B20"].value = "=ROUNDUP((1.3*B15), 0)-0.01" 
#     sheet["B21"].value =  "=1.219*B15"
#     workbook.save(filename = "item.xlsx")
#     workbook.save(filename = "eBay.xlsx")
    try:
        removeEbay(workbook)
    except ebaysdk.exception.ConnectionError:
        pass
    os.chdir(parentDir)


# URL = 'https://www.walmart.com/search/?query=tv'
# # # allPages = []
# # URL = "https://www.google.com"
# options = webdriver.ChromeOptions()
# options.add_argument(r"user-data-dir= C\Users\justa\AppData\Local\Google\Chrome\User Data\Default")
# # # options.add_argument('headless')
# options.add_argument('window-size=1300x600')
# driver = webdriver.Chrome(chrome_options = options)
# driver.get("https://target.com")
# driver.get(URL)
# links = driver.find_elements_by_class_name('product-title-link') #finds all links from product summaries on main page
# links[0].click()

# os.chdir("allItems")
# parentDir = os.getcwd()
# listDir = os.listdir()
# random.shuffle(listDir)
# body = ""
# for dir in listDir:
#     os.chdir(dir)
#     workbook = load_workbook("item.xlsx")
#     sheet = workbook[workbook.sheetnames[0]]
#     driver.get(sheet["B2"].value)
#     driver.implicitly_wait(random.randint(1, 8))
#     i = 0
#     while(driver.title == "Verify your identity"):
#         i += 1
#         if(i %1000 == 0):
#             print("Re")
        

#     print("")
#     print(sheet["B3"].value)
#     #price
#     try:
#         price = driver.find_element_by_class_name('price-characteristic').text
#         print("Price  = " + price + " old price  = " + sheet["B15"].value )
#         if(not(price == sheet["B15"].value)):
#             body += "Changed price for " + sheet["B3"].value + " from " + sheet["B15"].value + " to " + price + "\n"
#             sheet["B15"] = price
#             reviseEbay(workbook)
#     except:
#         exceptions.NoSuchElementException
#         pass

#     #Out of Stock
#     try:
#         stock = driver.find_element_by_class_name('display-block-xs').text
#         print(stock)
#         tryNext = False
#     except:
#         exceptions.NoSuchElementException
#         stock = "In Stock"
#         print(stock)
#         tryNext = True
    
#     if(tryNext):
#         print("try Next")
#         try:
#             stock = driver.find_element_by_class_name("prod-ProductOffer-urgencyMsg").text
#             print(stock)
#             stock = "Out of Stock"
#         except:
#             exceptions.NoSuchElementException
#             stock = "In Stock"
#             print(stock)
#     if(stock == "Out of Stock"):
#         body += "REMOVED " + sheet["B3"].value + "\n"
#         removeEbay(workbook)
#         print(os.getcwd())
#         os.chdir(parentDir)
#         shutil.rmtree(dir, ignore_errors=True)
#     os.chdir(parentDir)
# print(body)
# # page = WalmartPageInfo(URL, driver)
# # app = "JustinKl-test-PRD-1e65479d5-c023ee1b"
# # dev = "1823c83f-e9e2-467b-8a66-54bb1917eb6c"
# # cert = "PRD-e65479d53db2-e6ec-4713-b426-9429"
# # toke = "AgAAAA**AQAAAA**aAAAAA**829+Xw**nY+sHZ2PrBmdj6wVnY+sEZ2PrA2dj6MFlYGgCpGLogudj6x9nY+seQ**F3sGAA**AAMAAA**bwfjK8RqAGZYQ30CA3UapNkqeE4InpfhlfTM8dHhAPh/bF0RUAoKKBIVGHBNhN+EMdyvLOkVugCJtlo4FREbxKLh7aSE0NcNIwLbzjLJ5N9Ln1dzfmRo6pU9+AhHvygDxIRBJAbTunirpTjps+z4TghRo/ZkvevGAsmWe0SK9+0r6Z8p728AGvd47IeM1MrvD9dFJ/aUsQNeDR9gveNseY3Y7j+Wqa6CQiBH78hQcFgdoPkyvpGOe1FnslDQ5F2vy4H79n6yyN1cgbQVRJ78362OnY17BE7wdwmZOmiIW6ZEvXf8JYBAUAPeuoEUZo8AkZ6vKrX/PxlpWGVkS4RpIIe3lPUpaXccwzdz4n58XAVzSSq4e0q5srxXiiK7kkLXJGhTRWNacDFsnlyt6AFVgJ7N01Ic5fQ+zW516tF1KRpST/Tv41gTEFaM1BmWpZQXnPH6XhJtLQw07U+8g52qDNf+0PFRPGV2aIV1zTumWawuFw3zrLXPDiti9xf+XRZGIHjlNoxir05/oxA2jBai1PiLPYOm62Rlvv2tOdinH0AznDbMlX+DKyyLxXb5I31mZN+GjyNXzygmwfBOkrXlk2UzcQv070/yWXuq9EO91i7hMxaT4K+2Mk2VxBYERs5Bvbv8gBPFn8sDDMlN/nOvyYm6QiwQdF6XkTfpfJcG44qZbXfspnlHh2rasZQ5Ygmd7eVnzQFnBj+BqnoaFNmM2NlVl6ToNdAQu0ObtUggrYKw/FjqDrpWnvlalneGVHUK"

# # api = Connection(appid=app, devid=dev, certid=cert, token=toke, config_file=None, debug = True, escape_xml=True)
# # activeList = api.execute('GetMyeBaySelling', {"ActiveList": {"Include" :True} })

# # soup = BeautifulSoup(activeList.content, 'lxml')
# # ids = soup.find_all("itemid")
# # itemIDs = []
# # for i in ids:
# #     itemID = str(i)
# #     itemID = (((((itemID.replace("<itemid>", "")).replace("</itemid>", "")).replace("&amp;", "&")).replace("&lt;", "<")).replace("&gt;", ">")).replace("&apos;", "'").replace('&quot;', '"')
# #     itemIDs.append(itemID)
# #     print(itemID)
# # os.chdir("allItems")
# # parentDir = os.getcwd()
# # listDir = os.listdir()
# # print()
# # removed = []
# # for dir in listDir:
# #     os.chdir(dir)
# #     print(os.getcwd())
# #     print()
# #     workbook = load_workbook("item.xlsx")
# #     sheet = workbook[workbook.sheetnames[0]]
# #     redo = True
# #     beforeRemoved = True
# #     while(redo == True):
# #         try:
# #             for item in itemIDs:
# #                 print("if " + sheet["B23"].value + " excel == " + item)
# #                 if(sheet["B23"].value == item):
# #                     URL = sheet["B2"].value
# #                     if(URL == uBefore):
# #                         break
# #                     driver.get(URL)
# #                     page = WalmartPageInfo(URL, driver)
# #                     if(page.goodShipping == False):
# #                         removed.append(sheet["B3"].value)
# #                         beforeRemoved = True
# #                         removeEbay(workbook)
# #                         beforeRemoved = False
# #                         os.chdir("allItems")
# #                         shutil.rmtree(dir, ignore_errors=True)
# #                         body += str(i)+ ". Removed " + sheet["B3"].value + "\n"
# #             redo = False      
# #         except:
# #             print("Something Went Wrong")
# #             if(beforeRemoved):
# #                 if(driver.title == "Verify your identity"):
# #                     sendEmail("ReCaptcha","Wrong")
# #                     recaptcha = True
# #                 else:
# #                     recaptcha = False
# #                 some = 0
# #                 while(recaptcha):
# #                     if(some % 1000 == 0):
# #                         print("in Recaptcha")
# #                     if(not (driver.title == "Verify your identity")):
# #                         recaptcha = False
# #                         redo = True
# #                     some+=1
# #             else:
# #                 redo = False
# #         uBefore = URL
# #     os.chdir(parentDir)
# # print ("removed the following:")
# # for item in removed:
# #     print(item)