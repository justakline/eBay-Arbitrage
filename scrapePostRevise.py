import requests #Gets the information
import os
import smtplib
from bs4 import BeautifulSoup #Parses through information
from selenium import webdriver 
import time
from WalmartPageInfo import WalmartPageInfo
from selenium.common import exceptions  
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook
from openpyxl import load_workbook
import traceback
import urllib
import re
from pathlib import Path
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import shutil
import smtplib
import ebaysdk.trading

from addEbay import *
from removeEbay import *
from reviseEbay import *
from scrape_walmart import*

def checkListed(parentDir, driver):
    os.chdir("allItems")
    parentDir = os.getcwd()
    listDir = os.listdir()
    # random.shuffle(listDir)
    # listDir = listDir[28:]
    body = ""
    for dir in listDir:
        os.chdir(dir)
        workbook = load_workbook("item.xlsx")
        sheet = workbook[workbook.sheetnames[0]]
        try:
            driver.get(sheet["B2"].value)
            driver.implicitly_wait(random.randint(1, 8))
            i = 0
            try:
                if(driver.title == "Verify your identity"):
                    sendEmail("ReCaptcha","Wrong")
            except:
                exceptions.UnexpectedAlertPresentException
                pass
            again = True
            while(again == True):
                try:
                    if(driver.title == "Verify your identity"):
                        i += 1
                        if(i %1000 == 0):
                            print("Re")
                    else: 
                        again = False
                except:
                    exceptions.UnexpectedAlertPresentException
                    pass
                    
            while(driver.title == "Verify your identity"):
                i += 1
                if(i %1000 == 0):
                    print("Re")
        except:
            exceptions.TimeoutException
            print("Took too long for link to load")
            

        print("")
        print(sheet["B3"].value)
        #price
        try:
            price = driver.find_element_by_class_name('price-characteristic').text
            print("Price  = " + price + " old price  = " + sheet["B15"].value )
            if(not(price == sheet["B15"].value)):
                body += "Changed price for " + sheet["B3"].value + " from " + sheet["B15"].value + " to " + price + "\n"
                sheet["B15"] = price
                workbook.save("item.xlsx")
                workbook.save("eBay.xlsx")
                reviseEbay(workbook)

        except:
            exceptions.NoSuchElementException
            pass

        #Out of Stock
        try:
            stock = driver.find_element_by_class_name('display-block-xs').text
            print(stock)
            tryNext = False
        except:
            exceptions.NoSuchElementException
            stock = "In Stock"
            print(stock)
            tryNext = True
        
        if(tryNext):
            print("try Next")
            try:
                stock = driver.find_element_by_class_name("prod-ProductOffer-urgencyMsg").text
                print(stock)
                stock = "Out of Stock"
            except:
                exceptions.NoSuchElementException
                stock = "In Stock"
                print(stock)
        if(stock == "Out of Stock"):
            body += "REMOVED " + sheet["B3"].value + "\n"
            try:
                removeEbay(workbook)
            except:
                pass
            print(os.getcwd())
            os.chdir(parentDir)
            shutil.rmtree(dir, ignore_errors=True)
        os.chdir(parentDir)
    print(body)
    driver.quit()

# def getAllListedItemLinks(parentDir):
#     os.chdir("allItems")
#     parentDir = os.getcwd()
#     listDir = os.listdir()
#     random.shuffle(listDir)
#     listed = []
#     for dir in listDir:
#         os.chdir(dir)
#         workbook = load_workbook("item.xlsx")
#         sheet = workbook[workbook.sheetnames[0]]
#         listed.append(sheet["B2"].value)
#         os.chdir(parentDir)
#     os.chdir(parentDir)
#     return listed



app = "JustinKl-test-PRD-1e65479d5-c023ee1b"
dev = "1823c83f-e9e2-467b-8a66-54bb1917eb6c"
cert = "PRD-e65479d53db2-e6ec-4713-b426-9429"
toke = "AgAAAA**AQAAAA**aAAAAA**829+Xw**nY+sHZ2PrBmdj6wVnY+sEZ2PrA2dj6MFlYGgCpGLogudj6x9nY+seQ**F3sGAA**AAMAAA**bwfjK8RqAGZYQ30CA3UapNkqeE4InpfhlfTM8dHhAPh/bF0RUAoKKBIVGHBNhN+EMdyvLOkVugCJtlo4FREbxKLh7aSE0NcNIwLbzjLJ5N9Ln1dzfmRo6pU9+AhHvygDxIRBJAbTunirpTjps+z4TghRo/ZkvevGAsmWe0SK9+0r6Z8p728AGvd47IeM1MrvD9dFJ/aUsQNeDR9gveNseY3Y7j+Wqa6CQiBH78hQcFgdoPkyvpGOe1FnslDQ5F2vy4H79n6yyN1cgbQVRJ78362OnY17BE7wdwmZOmiIW6ZEvXf8JYBAUAPeuoEUZo8AkZ6vKrX/PxlpWGVkS4RpIIe3lPUpaXccwzdz4n58XAVzSSq4e0q5srxXiiK7kkLXJGhTRWNacDFsnlyt6AFVgJ7N01Ic5fQ+zW516tF1KRpST/Tv41gTEFaM1BmWpZQXnPH6XhJtLQw07U+8g52qDNf+0PFRPGV2aIV1zTumWawuFw3zrLXPDiti9xf+XRZGIHjlNoxir05/oxA2jBai1PiLPYOm62Rlvv2tOdinH0AznDbMlX+DKyyLxXb5I31mZN+GjyNXzygmwfBOkrXlk2UzcQv070/yWXuq9EO91i7hMxaT4K+2Mk2VxBYERs5Bvbv8gBPFn8sDDMlN/nOvyYm6QiwQdF6XkTfpfJcG44qZbXfspnlHh2rasZQ5Ygmd7eVnzQFnBj+BqnoaFNmM2NlVl6ToNdAQu0ObtUggrYKw/FjqDrpWnvlalneGVHUK"

api = Connection(appid=app, devid=dev, certid=cert, token=toke, config_file=None, debug = True, escape_xml=True)
passive = os.getcwd()
# listed = getAllListedItemLinks(passive)
options = webdriver.ChromeOptions()
options.add_argument(r"user-data-dir= C\Users\justa\AppData\Local\Google\Chrome\User Data\Default")
options.add_argument('window-size=1200x600')

driver = webdriver.Chrome(chrome_options = options)
# checkListed(passive, driver)
driver.get("https://walmart.com")
driver.implicitly_wait(4)
# URL = 'https://www.walmart.com/search/?query=TVs' #Cheaper TVS
# scrape_walmart(driver, URL)

URL = 'https://www.walmart.com/search/?cat_id=0&facet=price%3A%24400+-+%24450%7C%7Cprice%3A%24450+-+%24500%7C%7Cprice%3A%24500+-+%24700&query=tv' #Expensive TVS
scrape_walmart(driver, URL)

os.chdir(passive)
os.chdir("allItems")
parentDir = os.getcwd()
listDir = os.listdir()
activeList = api.execute('GetMyeBaySelling', {"ActiveList": {"Include" :True} })
soup = BeautifulSoup(activeList.content, 'lxml')
tits = soup.find_all("title")
titles = []
ides = soup.find_all("itemid")
ids = []
for i in range(len(tits)):
    tit = str(tits[i]) + ""
    tit = tit.replace("<title>", "")
    tit = tit.replace("</title>", "")
    tit = (tit).replace("&amp;", "&")
    tit = (((tit.replace("&lt;", "<")).replace("&gt;", ">")).replace("&apos;", "'")).replace('&quot;', '"')
    titles.append(tit)
    ins = str(ides[i]) + ""
    ins = ((((((ins.replace("<itemid>", "")).replace("</title>", "")).replace("&amp;", "&")).replace("&lt;", "<")).replace("&gt;", ">")).replace("&apos;", "'")).replace('&quot;', '"')
    ids.append(ins)

i = 0
body = ""
for dir in listDir:
    os.chdir(dir)
    workbook = load_workbook("item.xlsx")
    sheet = workbook[workbook.sheetnames[0]]
    try:
        if(sheet["B5"].value == "Out of Stock"):
            removeEbay(workbook)
            os.chdir(dir)
            os.chdir(parentDir)
            shutil.rmtree(dir, ignore_errors=True)
            body += str(i)+ ". Removed " + sheet["B3"].value + "\n"
        elif(sheet["B24"].value == "True"):
            sheet["B24"].value = "False"
            body += str(i)+ ". Revised " + sheet["B3"].value + " to the price of $" + getPrice(sheet) + "\n"
            reviseEbay(workbook)
        elif(sheet["B3"].value in titles):
            body+= str(i) + ". " + sheet["B3"].value + " had nothing happen \n"
            pass
        elif(sheet["B23"].value in ids):
            body+= str(i) + ". " + sheet["B3"].value + " had nothing happen \n"
            pass
        else:
            addEbay(workbook)
            body += str(i)+ ". Added " + sheet["B3"].value + " to the price of $" + getPrice(sheet) + " and the ItemID = " + sheet["B23"].value+ "\n"
        i+=1
    except ebaysdk.exception.ConnectionError:
        print("error")
        pass
    os.chdir(parentDir)
print()
print("Finished")
print()
print(body)
if(not (body == "")): 
    sendEmail("Ebay Postings, Removals, and Revisions ", body)
else:
    sendEmail("Nothing happened this round", "nothing")

